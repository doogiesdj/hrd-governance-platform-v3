import json
import openpyxl
from pathlib import Path

_SCRIPTS_DIR = Path(__file__).parent
_REPO_ROOT = _SCRIPTS_DIR.parent
EXCEL_PATH = str(_REPO_ROOT / "src" / "HRD_CodeMapping_v2.xlsx")
OUTPUT_PATH = str(_REPO_ROOT / "data" / "code_reference.json")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

result = {}
sheet_meta = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0] if len(row) > 0 else None
        name = row[1] if len(row) > 1 else None
        if code and name:
            rows.append({"code": str(code).strip(), "name": str(name).strip()})
    result[sheet_name] = rows
    sheet_meta.append({"sheet": sheet_name, "count": len(rows)})
    print(f"Sheet '{sheet_name}': {len(rows)} rows", flush=True)

output = {"sheets": result, "meta": sheet_meta}

Path(OUTPUT_PATH).parent.mkdir(parents=True, exist_ok=True)
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\nDone. Written to {OUTPUT_PATH}")

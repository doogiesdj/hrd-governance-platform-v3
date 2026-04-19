"""
Full ontology extractor: merges Excel + RDF into enriched ontology.json
"""
import xml.etree.ElementTree as ET
import openpyxl
import json
import sys
import io
import re
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

_SCRIPTS_DIR = Path(__file__).parent
_REPO_ROOT = _SCRIPTS_DIR.parent
_EXCEL_PATH = str(_REPO_ROOT / "src" / "HRD_CodeMapping_v2.xlsx")
_RDF_PATH = str(_REPO_ROOT / "src" / "HRD_Governance_Extended_v2.rdf")
_ONTOLOGY_JSON = str(_REPO_ROOT / "data" / "ontology.json")
_ONTOLOGY_JS = str(_REPO_ROOT / "js" / "ontology_data.js")

RDF_NS = 'http://www.w3.org/1999/02/22-rdf-syntax-ns#'
OWL_NS = 'http://www.w3.org/2002/07/owl#'
RDFS_NS = 'http://www.w3.org/2000/01/rdf-schema#'
HRD_NS = 'http://www.human-resource.go.kr/ontology/hrd#'

def rdf_prop(tag):
    return f'{{{HRD_NS}}}{tag}'

# ─── Read Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)

def get_rows(sheet_name):
    ws = wb[sheet_name]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        result.append(dict(zip(header, row)))
    return result

xl_strategies   = {r['코드']: r for r in get_rows('국가전략(NationalStrategy)')}
xl_policies     = {r['코드']: r for r in get_rows('정책(Policy)')}
xl_org_budgets  = {r['코드']: r for r in get_rows('기관배정예산(OrgBudget)')}
xl_pol_budgets  = {r['코드']: r for r in get_rows('정책별예산(PolicyBudget)')}
xl_orgs         = {r['코드']: r for r in get_rows('기관(Organization)')}
xl_programs     = {r['코드']: r for r in get_rows('교육프로그램(EducationProgram)')}
xl_tgroups      = {r['코드']: r for r in get_rows('대상집단(TargetGroup)')}
xl_comps        = {r['코드']: r for r in get_rows('역량(Competency)')}
xl_outcomes     = {r['코드']: r for r in get_rows('성과(Outcome)')}
xl_comp_cats    = {r['코드']: r for r in get_rows('역량분류(CompetencyCategory)')}
xl_persons      = {r['코드']: r for r in get_rows('인원(Person)')}
xl_comp_assess  = {r['코드']: r for r in get_rows('역량평가(CompetencyAssessment)')}
xl_benefits     = {r['코드']: r for r in get_rows('혜택(Benefit)')}
xl_pol_parts    = {r['코드']: r for r in get_rows('정책참여(PolicyParticipation)')}
xl_recommends   = {r['코드']: r for r in get_rows('추천(Recommendation)')}
xl_comp_gaps    = {r['코드']: r for r in get_rows('역량갭(CompetencyGap)')}
xl_enrollments  = {r['코드']: r for r in get_rows('프로그램등록(ProgramEnrollment)')}

def parse_codes(s):
    if not s:
        return []
    return [c.strip() for c in str(s).split(',') if c.strip()]

# ─── Read RDF ─────────────────────────────────────────────────────────────────
tree = ET.parse(_RDF_PATH)
root = tree.getroot()

rdf_data = {}  # id -> {prop -> value or list}

for ind in root.iter(f'{{{OWL_NS}}}NamedIndividual'):
    about = ind.get(f'{{{RDF_NS}}}about', '')
    name = about.split('#')[-1] if '#' in about else about
    props = {}
    for child in ind:
        raw_tag = child.tag
        ns = raw_tag.split('}')[0].lstrip('{') if '}' in raw_tag else ''
        tag = raw_tag.split('}')[-1]
        resource = child.get(f'{{{RDF_NS}}}resource', '')
        ref = resource.split('#')[-1] if '#' in resource else ''
        text = (child.text or '').strip()
        val = ref if ref else text
        if not val:
            continue
        if tag in props:
            existing = props[tag]
            if isinstance(existing, list):
                existing.append(val)
            else:
                props[tag] = [existing, val]
        else:
            props[tag] = val
    rdf_data[name] = props

# Also parse rdf:Description elements (used for additional property assertions)
def _merge_props(target_dict, name, child_iter):
    for child in child_iter:
        tag = child.tag.split('}')[-1]
        resource = child.get(f'{{{RDF_NS}}}resource', '')
        ref = resource.split('#')[-1] if '#' in resource else ''
        text = (child.text or '').strip()
        val = ref if ref else text
        if not val:
            continue
        entry = target_dict.setdefault(name, {})
        if tag in entry:
            existing = entry[tag]
            if isinstance(existing, list):
                existing.append(val)
            else:
                entry[tag] = [existing, val]
        else:
            entry[tag] = val

for desc in root.iter(f'{{{RDF_NS}}}Description'):
    about = desc.get(f'{{{RDF_NS}}}about', '')
    name = about.split('#')[-1] if '#' in about else about
    if name:
        _merge_props(rdf_data, name, desc)

def rdf_get(name, prop, default=None):
    d = rdf_data.get(name, {})
    v = d.get(prop, default)
    return v

def rdf_list(name, prop):
    v = rdf_get(name, prop)
    if v is None:
        return []
    return v if isinstance(v, list) else [v]

# ─── Build reverse indexes ─────────────────────────────────────────────────────
program_to_strategy = {}
strategy_to_policies = {}

# Program → strategy from Excel program.정렬전략
for prog_id, prog in xl_programs.items():
    strat = prog.get('정렬전략')
    if strat and strat in xl_strategies:
        program_to_strategy.setdefault(prog_id, set()).add(strat)

# Policy → strategy via PolicyBudget (관련정책 → 관련전략)
for pb_id, pb in xl_pol_budgets.items():
    strat_code = pb.get('관련전략', '').strip()
    pol_codes_raw = pb.get('관련정책', '')
    if not strat_code or strat_code not in xl_strategies:
        continue
    pol_codes = parse_codes(str(pol_codes_raw)) if pol_codes_raw else []
    for pc in pol_codes:
        if pc in xl_policies:
            strategy_to_policies.setdefault(strat_code, set()).add(pc)

# Build strategy→programs from program_to_strategy reverse
strategy_to_programs = {}
for prog_id, strats in program_to_strategy.items():
    for s in strats:
        strategy_to_programs.setdefault(s, set()).add(prog_id)

# Build strategy→budgets from OrgBudget Excel (관련전략)
strategy_to_budgets = {}
for bud_id, bud in xl_org_budgets.items():
    strat = bud.get('관련전략')
    if strat and strat in xl_strategies:
        strategy_to_budgets.setdefault(strat, []).append(bud_id)

# Build policy→organizations from PolicyBudget (관련정책 → 관리기관코드)
policy_to_orgs = {}
for pb_id, pb in xl_pol_budgets.items():
    pol_codes = parse_codes(str(pb.get('관련정책', '') or ''))
    org_code = pb.get('관리기관코드', '')
    for pc in pol_codes:
        if org_code:
            policy_to_orgs.setdefault(pc, set()).add(org_code)

# Build policy→policy_budgets map
policy_to_pol_budgets = {}
for pb_id, pb in xl_pol_budgets.items():
    pol_codes = parse_codes(str(pb.get('관련정책', '') or ''))
    for pc in pol_codes:
        policy_to_pol_budgets.setdefault(pc, []).append(pb_id)

# Build org→programs map
org_to_programs = {}
for prog_id, prog in xl_programs.items():
    oid = prog.get('운영기관', '')
    if oid:
        org_to_programs.setdefault(oid, []).append(prog_id)

# Build org→budgets map (from OrgBudget)
org_to_budgets = {}
for bud_id, bud in xl_org_budgets.items():
    oid = bud.get('관리기관코드', '')
    if oid:
        org_to_budgets.setdefault(oid, []).append(bud_id)

# ─── Build strategy details ────────────────────────────────────────────────────
def amt_to_int(s):
    if not s:
        return 0
    return int(str(s).replace(',', ''))

def format_won(n):
    if n >= 1_000_000_000_000:
        return f'{n/1_000_000_000_000:.1f}조원'
    if n >= 100_000_000:
        return f'{n/100_000_000:.0f}억원'
    if n >= 10_000:
        return f'{n/10_000:.0f}만원'
    return f'{n:,}원'

strategies_out = []
for strat_id, strat_xl in xl_strategies.items():
    rdf_props = rdf_data.get(strat_id, {})

    # Policies linked to this strategy
    linked_policy_ids = list(strategy_to_policies.get(strat_id, set()))
    linked_policies = []
    for pid in linked_policy_ids:
        p = xl_policies.get(pid)
        if p:
            linked_policies.append({'id': pid, 'name': p['한글명'], 'en': p['영문명']})

    # Programs
    linked_prog_ids = list(strategy_to_programs.get(strat_id, set()))
    linked_programs = []
    for pid in linked_prog_ids[:20]:
        p = xl_programs.get(pid)
        if p:
            tg = xl_tgroups.get(p.get('대상집단', ''), {})
            org = xl_orgs.get(p.get('운영기관', ''), {})
            comp = xl_comps.get(p.get('목표역량', ''), {})
            linked_programs.append({
                'id': pid,
                'name': p['한글명'],
                'en': p['영문명'],
                'targetGroup': tg.get('한글명', p.get('대상집단', '')),
                'org': org.get('한글명', p.get('운영기관', '')),
                'competency': comp.get('한글명', p.get('목표역량', '')),
                'hours': p.get('교육시간', ''),
            })

    # Budget
    bud_ids = strategy_to_budgets.get(strat_id, [])
    total_budget = 0
    budget_items = []
    for bid in bud_ids:
        b = xl_org_budgets.get(bid, {})
        amt = amt_to_int(b.get('예산총액(원)', 0))
        org = xl_orgs.get(b.get('관리기관코드', ''), {})
        total_budget += amt
        budget_items.append({
            'id': bid,
            'name': b.get('한글명', ''),
            'amount': amt,
            'amountStr': format_won(amt),
            'org': org.get('한글명', b.get('관리기관코드', '')),
            'year': b.get('회계연도', ''),
        })

    # Implementing orgs (from programs)
    org_set = {}
    for pid in linked_prog_ids:
        p = xl_programs.get(pid, {})
        oid = p.get('운영기관')
        if oid and oid not in org_set:
            org = xl_orgs.get(oid, {})
            org_set[oid] = {
                'id': oid,
                'name': org.get('한글명', oid),
                'abbr': org.get('약칭', ''),
                'role': org.get('HRD역할', ''),
            }
    # Also add budget-managing orgs
    for bid in bud_ids:
        b = xl_org_budgets.get(bid, {})
        oid = b.get('관리기관코드')
        if oid and oid not in org_set:
            org = xl_orgs.get(oid, {})
            org_set[oid] = {
                'id': oid,
                'name': org.get('한글명', oid),
                'abbr': org.get('약칭', ''),
                'role': org.get('HRD역할', ''),
            }
    implementing_orgs = list(org_set.values())

    # Target groups (from programs)
    tgroup_set = {}
    for pid in linked_prog_ids:
        p = xl_programs.get(pid, {})
        tgid = p.get('대상집단')
        if tgid and tgid not in tgroup_set:
            tg = xl_tgroups.get(tgid, {})
            tgroup_set[tgid] = tg.get('한글명', tgid)
    target_groups_list = list(tgroup_set.values())

    # Competencies (from programs)
    comp_set = {}
    for pid in linked_prog_ids:
        p = xl_programs.get(pid, {})
        cid = p.get('목표역량')
        if cid and cid not in comp_set:
            comp = xl_comps.get(cid, {})
            cat_id = comp.get('역량분류', '')
            cat = xl_comp_cats.get(cat_id, {})
            comp_set[cid] = {
                'id': cid,
                'name': comp.get('한글명', cid),
                'en': comp.get('영문명', ''),
                'category': cat.get('한글명', cat_id),
            }
    competencies_list = list(comp_set.values())

    # Description from RDF
    desc_ko = rdf_props.get('strategyDescription', '')
    desc_en = rdf_props.get('strategyDescriptionEn', '')

    # Performance goals from Outcomes linked to policies of this strategy
    perf_goals = []
    for pol in linked_policies[:5]:
        pol_xl = xl_policies.get(pol['id'], {})
        out_codes = parse_codes(pol_xl.get('지원성과', ''))
        for oc in out_codes[:3]:
            out = xl_outcomes.get(oc, {})
            if out:
                perf_goals.append({
                    'code': oc,
                    'name': out.get('한글명', oc),
                    'type': out.get('성과유형', ''),
                    'value': out.get('성과값', ''),
                })

    strategies_out.append({
        'id': strat_id,
        'name': strat_xl['한글명'],
        'en': strat_xl['영문명'],
        'priority': strat_xl.get('우선순위', ''),
        'description': desc_ko or f'{strat_xl["한글명"]}은 국가 인적자원 개발의 핵심 전략으로, 관련 정책 및 교육 프로그램을 통해 추진됩니다.',
        'descriptionEn': desc_en or '',
        'type': 'NationalStrategy',
        'policies': linked_policies,
        'programs': linked_programs,
        'budgets': budget_items,
        'totalBudget': total_budget,
        'totalBudgetStr': format_won(total_budget) if total_budget else '미배정',
        'implementingOrgs': implementing_orgs,
        'targetGroups': target_groups_list,
        'competencies': competencies_list,
        'performanceGoals': perf_goals,
        'policyCount': len(linked_policies),
        'programCount': len(linked_prog_ids),
    })

# Sort by priority
strategies_out.sort(key=lambda s: int(str(s['priority'])) if str(s.get('priority',99)).isdigit() else 99)

# ─── Build other entities ──────────────────────────────────────────────────────
# Reverse map: policy → strategy (from strategy_to_policies)
policy_to_strategy_map = {}
for strat_id, pol_set in strategy_to_policies.items():
    for pid in pol_set:
        policy_to_strategy_map[pid] = strat_id

def budget_scale(amt):
    if amt >= 10_000_000_000:  # 100억+
        return '대형'
    if amt >= 1_000_000_000:   # 10억~100억
        return '중형'
    if amt > 0:
        return '소형'
    return '미배정'

policies_out = []
for pol_id, pol in xl_policies.items():
    comp_codes = parse_codes(pol.get('지원역량', ''))
    comps = []
    comp_cats = {}
    for c in comp_codes:
        if not c:
            continue
        comp_xl = xl_comps.get(c, {})
        cat_id = comp_xl.get('역량분류', '')
        cat_xl = xl_comp_cats.get(cat_id, {})
        cat_name = cat_xl.get('한글명', cat_id)
        comps.append({'id': c, 'name': comp_xl.get('한글명', c), 'category': cat_name})
        if cat_name:
            comp_cats[cat_name] = True

    out_codes = parse_codes(pol.get('지원성과', ''))
    perf_goals = []
    for oc in out_codes:
        out = xl_outcomes.get(oc, {})
        if out:
            perf_goals.append({
                'code': oc,
                'name': out.get('한글명', oc).split('/')[0].strip(),
                'type': out.get('성과유형', ''),
                'value': out.get('성과값', ''),
            })

    bud = xl_pol_budgets.get(pol.get('배정예산', ''), {})
    amt = amt_to_int(bud.get('예산총액(원)', 0))
    org_code = bud.get('관리기관코드', '')
    org = xl_orgs.get(org_code, {})

    strat_code = policy_to_strategy_map.get(pol_id, '')
    strat_name = xl_strategies.get(strat_code, {}).get('한글명', '') if strat_code else ''

    related_org_ids = list(policy_to_orgs.get(pol_id, set()))
    related_orgs = []
    for oid in related_org_ids:
        o = xl_orgs.get(oid, {})
        related_orgs.append({'id': oid, 'name': o.get('한글명', oid), 'abbr': o.get('약칭', '')})

    related_bud_ids = policy_to_pol_budgets.get(pol_id, [])
    related_budgets = []
    for bid in related_bud_ids:
        b = xl_pol_budgets.get(bid, {})
        b_amt = amt_to_int(b.get('예산총액(원)', 0))
        related_budgets.append({'id': bid, 'name': b.get('한글명', bid), 'amount': b_amt, 'amountStr': format_won(b_amt) if b_amt else '미배정'})

    policies_out.append({
        'id': pol_id,
        'name': pol['한글명'].split('/')[0].strip(),
        'en': pol['영문명'],
        'type': 'PublicPolicy',
        'budgetId': pol.get('배정예산', ''),
        'budgetAmount': amt,
        'budgetAmountStr': format_won(amt) if amt else '미배정',
        'budgetScale': budget_scale(amt),
        'managingOrg': org.get('한글명', ''),
        'managingOrgAbbr': org.get('약칭', ''),
        'managingOrgType': org.get('기관유형', ''),
        'relatedStrategy': strat_code,
        'relatedStrategyName': strat_name,
        'relatedOrgs': related_orgs,
        'relatedBudgets': related_budgets,
        'competencies': comps,
        'competencyCategories': list(comp_cats.keys()),
        'performanceGoals': perf_goals,
    })

orgs_out = []
for org_id, org in xl_orgs.items():
    prog_ids = org_to_programs.get(org_id, [])
    prog_list = []
    for pid in prog_ids[:10]:
        p = xl_programs.get(pid, {})
        prog_list.append({'id': pid, 'name': p.get('한글명', pid)})

    bud_ids = org_to_budgets.get(org_id, [])
    bud_list = []
    total_bud = 0
    for bid in bud_ids:
        b = xl_org_budgets.get(bid, {})
        b_amt = amt_to_int(b.get('예산총액(원)', 0))
        total_bud += b_amt
        bud_list.append({'id': bid, 'name': b.get('한글명', bid), 'amount': b_amt, 'amountStr': format_won(b_amt) if b_amt else '미배정'})

    orgs_out.append({
        'id': org_id,
        'name': org['한글명'],
        'en': org['영문명'],
        'abbr': org.get('약칭', ''),
        'type': org.get('기관유형', ''),
        'role': org.get('HRD역할', ''),
        'dept': org.get('담당부서', ''),
        'mainTel': org.get('대표전화', ''),
        'directTel': org.get('담당자전화', ''),
        'parentOrgId': org.get('상위기관코드', ''),
        'programs': prog_list,
        'programCount': len(prog_ids),
        'budgets': bud_list,
        'totalBudget': total_bud,
        'totalBudgetStr': format_won(total_bud) if total_bud else '미배정',
    })

budgets_out = []
total_strat_budget = 0
for bud_id, bud in xl_org_budgets.items():
    amt = amt_to_int(bud.get('예산총액(원)', 0))
    total_strat_budget += amt
    org = xl_orgs.get(bud.get('관리기관코드', ''), {})
    budgets_out.append({
        'id': bud_id,
        'name': bud.get('한글명', ''),
        'en': bud.get('영문명', ''),
        'amount': amt,
        'amountStr': format_won(amt),
        'fiscalYear': bud.get('회계연도', ''),
        'managingOrgId': bud.get('관리기관코드', ''),
        'managingOrg': org.get('한글명', bud.get('관리기관코드', '')),
        'relatedStrategy': bud.get('관련전략', ''),
        'budgetType': 'OrgBudget',
        'type': 'Budget',
    })

for bud_id, bud in xl_pol_budgets.items():
    amt = amt_to_int(bud.get('예산총액(원)', 0))
    total_strat_budget += amt
    org = xl_orgs.get(bud.get('관리기관코드', ''), {})
    pol_codes = parse_codes(str(bud.get('관련정책', '') or ''))
    pol_names = [xl_policies.get(pc, {}).get('한글명', pc) for pc in pol_codes if pc]
    budgets_out.append({
        'id': bud_id,
        'name': bud.get('한글명', ''),
        'en': bud.get('영문명', ''),
        'amount': amt,
        'amountStr': format_won(amt),
        'fiscalYear': bud.get('회계연도', ''),
        'managingOrgId': bud.get('관리기관코드', ''),
        'managingOrg': org.get('한글명', bud.get('관리기관코드', '')),
        'relatedStrategy': bud.get('관련전략', ''),
        'relatedPolicies': pol_codes,
        'relatedPolicyNames': pol_names,
        'budgetType': 'PolicyBudget',
        'type': 'Budget',
    })

programs_out = []
for prog_id, prog in xl_programs.items():
    tg = xl_tgroups.get(prog.get('대상집단', ''), {})
    org = xl_orgs.get(prog.get('운영기관', ''), {})
    comp = xl_comps.get(prog.get('목표역량', ''), {})
    cat = xl_comp_cats.get(comp.get('역량분류', ''), {}) if comp else {}
    budget_code = str(prog.get('재원', '') or '')
    rdf_type = rdf_get(prog_id, 'type', 'EducationProgram')
    prog_subclass = rdf_type if rdf_type in (
        'K_DigitalTraining', 'DegreeCourse', 'JobRetraining', 'OnlineModule', 'WorkshopSeminar'
    ) else 'EducationProgram'
    programs_out.append({
        'id': prog_id,
        'name': prog['한글명'],
        'en': prog['영문명'],
        'type': prog_subclass,
        'targetGroupId': prog.get('대상집단', ''),
        'targetGroup': tg.get('한글명', prog.get('대상집단', '')),
        'targetGroupEn': tg.get('영문명', ''),
        'orgId': prog.get('운영기관', ''),
        'org': org.get('한글명', prog.get('운영기관', '')),
        'orgAbbr': org.get('약칭', ''),
        'alignedStrategy': prog.get('정렬전략', ''),
        'competency': comp.get('한글명', prog.get('목표역량', '')),
        'competencyEn': comp.get('영문명', ''),
        'competencyCategory': cat.get('한글명', ''),
        'targetCompetencyId': prog.get('목표역량', ''),
        'fieldCatCode': str(prog.get('역량분야', '') or ''),
        'relatedCompetencyIds': rdf_list(prog_id, 'developsCompetency'),
        'budgetCode': budget_code,
        'hours': prog.get('교육시간', ''),
    })

CATCODE_TO_PROTEGE = {
    'COMPCAT_Leadership':            ('SoftSkill', 'Interpersonal', 'Leadership'),
    'COMPCAT_Strategic':             ('HardSkill', 'Business_Admin', 'Management'),
    'COMPCAT_Capability':            ('HardSkill', 'Business_Admin', 'Management'),
    'COMPCAT_Marketing_Strategy':    ('HardSkill', 'Business_Admin', 'Marketing_Strategy'),
    'COMPCAT_AI_and_Infrastructure': ('HardSkill', 'ICT_Dev', 'AI_and_Infrastructure'),
    'COMPCAT_Software':              ('HardSkill', 'ICT_Dev', 'Software'),
    'COMPCAT_Technical':             ('HardSkill', 'ICT_Dev', 'Software'),
    'COMPCAT_Design':                ('HardSkill', 'ICT_Dev', 'Software'),
    'COMPCAT_BioHealth':             ('HardSkill', 'Industrial_Tech', 'Manufacturing'),
    'COMPCAT_Mobility':              ('HardSkill', 'Industrial_Tech', 'Manufacturing'),
    'COMPCAT_Semiconductor_Design':  ('HardSkill', 'Industrial_Tech', 'Manufacturing'),
    'COMPCAT_Shipbuilding':          ('HardSkill', 'Industrial_Tech', 'Manufacturing'),
    'COMPCAT_Space_Aerospace':       ('HardSkill', 'Industrial_Tech', 'Manufacturing'),
    'COMPCAT_Green_Tech':            ('HardSkill', 'Industrial_Tech', 'Manufacturing'),
    'COMPCAT_Defense_Industry':      ('HardSkill', 'Industrial_Tech', 'Manufacturing'),
    'COMPCAT_AdminLaw':              ('HardSkill', 'Industrial_Tech', 'Service_Public'),
    'COMPCAT_Communication':         ('Literacy', 'Basic_Academic', 'Language_and_Math'),
    'COMPCAT_Values':                ('Literacy', 'Civic_Literacy', 'Social_Value'),
    'COMPCAT_Data_Fluency':          ('Literacy', 'Digital_Literacy', 'Data_Fluency'),
    'COMPCAT_Tech_Security':         ('Literacy', 'Digital_Literacy', 'Tech_Security'),
    'COMPCAT_Interpersonal':         ('SoftSkill', 'Interpersonal', 'Collaboration'),
    'COMPCAT_Thinking':              ('SoftSkill', 'Problem_Solving', 'Thinking_Skill'),
    'COMPCAT_Personal':              ('SoftSkill', 'Self_Management', 'Adaptability'),
    'COMPCAT_Learning':              ('SoftSkill', 'Self_Management', 'Adaptability'),
}

COMP_OVERRIDES = {
    'COMP_35': ('HardSkill', 'Business_Admin', 'Management'),
    'COMP_36': ('HardSkill', 'Business_Admin', 'Management'),
    'COMP_37': ('HardSkill', 'Business_Admin', 'Management'),
    'COMP_74': ('Literacy', 'Basic_Academic', 'Language_and_Math'),
    'COMP_75': ('SoftSkill', 'Self_Management', 'Adaptability'),
}

comps_out = []
for comp_id, comp in xl_comps.items():
    cat = xl_comp_cats.get(comp.get('역량분류', ''), {})
    cat_code = comp.get('역량분류', '')
    protege = COMP_OVERRIDES.get(comp_id, CATCODE_TO_PROTEGE.get(cat_code, ('', '', '')))
    comps_out.append({
        'id': comp_id,
        'name': comp['한글명'],
        'en': comp['영문명'],
        'type': 'Competency',
        'category': cat.get('한글명', cat_code),
        'categoryEn': cat.get('영문명', ''),
        'catCode': cat_code,
        'class': protege[0],
        'level2': protege[1],
        'level3': protege[2],
    })

tgroups_out = [
    {'id': tid, 'name': tg['한글명'], 'en': tg['영문명']}
    for tid, tg in xl_tgroups.items()
]

# ─── New entity types ──────────────────────────────────────────────────────────
persons_out = []
for pid, person in xl_persons.items():
    tg = xl_tgroups.get(str(person.get('대상집단', '') or ''), {})
    comp = xl_comps.get(str(person.get('보유역량', '') or ''), {})
    strat = xl_strategies.get(str(person.get('기여전략', '') or ''), {})
    persons_out.append({
        'id': pid,
        'name': person['한글명'],
        'en': person.get('영문명', ''),
        'type': 'Person',
        'targetGroupId': str(person.get('대상집단', '') or ''),
        'targetGroup': tg.get('한글명', str(person.get('대상집단', '') or '')),
        'employmentStatus': str(person.get('고용상태', '') or ''),
        'contributingStrategyId': str(person.get('기여전략', '') or ''),
        'contributingStrategy': strat.get('한글명', str(person.get('기여전략', '') or '')),
        'competencyId': str(person.get('보유역량', '') or ''),
        'competency': comp.get('한글명', str(person.get('보유역량', '') or '')),
        'rank': str(person.get('직급', '') or ''),
        'department': str(person.get('부서', '') or ''),
    })

competencyAssessments_out = []
for aid, assess in xl_comp_assess.items():
    comp = xl_comps.get(str(assess.get('평가역량', '') or ''), {})
    competencyAssessments_out.append({
        'id': aid,
        'name': assess['한글명'],
        'en': assess.get('영문명', ''),
        'type': 'CompetencyAssessment',
        'competencyId': str(assess.get('평가역량', '') or ''),
        'competency': comp.get('한글명', str(assess.get('평가역량', '') or '')),
        'priorScore': assess.get('사전점수'),
        'postScore': assess.get('사후점수'),
        'targetScore': assess.get('목표점수'),
    })

benefits_out = []
for bid, benefit in xl_benefits.items():
    benefits_out.append({
        'id': bid,
        'name': benefit['한글명'],
        'en': benefit.get('영문명', ''),
        'type': 'Benefit',
        'benefitType': str(benefit.get('혜택유형', '') or ''),
        'benefitValue': str(benefit.get('혜택가치', '') or ''),
    })

outcomes_out = []
for oid, outcome in xl_outcomes.items():
    comp = xl_comps.get(str(outcome.get('향상역량', '') or ''), {})
    pol = xl_policies.get(str(outcome.get('연계정책', '') or ''), {})
    outcomes_out.append({
        'id': oid,
        'name': outcome['한글명'],
        'en': outcome.get('영문명', ''),
        'type': 'Outcome',
        'outcomeType': str(outcome.get('성과유형', '') or ''),
        'outcomeValue': str(outcome.get('성과값', '') or ''),
        'achievedCount': outcome.get('달성인원'),
        'improvedCompetencyId': str(outcome.get('향상역량', '') or ''),
        'improvedCompetency': comp.get('한글명', str(outcome.get('향상역량', '') or '')),
        'linkedPolicyId': str(outcome.get('연계정책', '') or ''),
        'linkedPolicy': pol.get('한글명', str(outcome.get('연계정책', '') or '')),
    })

policyParticipations_out = []
for ppid, pp in xl_pol_parts.items():
    benefit_codes = parse_codes(str(pp.get('혜택', '') or ''))
    benefit_names = [xl_benefits.get(bc, {}).get('한글명', bc) for bc in benefit_codes if bc]
    outcome_codes = parse_codes(str(pp.get('성과', '') or ''))
    outcome_names = [xl_outcomes.get(oc, {}).get('한글명', oc) for oc in outcome_codes if oc]
    policyParticipations_out.append({
        'id': ppid,
        'name': pp['한글명'],
        'en': pp.get('영문명', ''),
        'type': 'PolicyParticipation',
        'participantCount': pp.get('참여인원'),
        'benefitIds': benefit_codes,
        'benefits': benefit_names,
        'outcomeIds': outcome_codes,
        'outcomes': outcome_names,
    })

recommendations_out = []
for rid, rec in xl_recommends.items():
    prog_codes = parse_codes(str(rec.get('추천프로그램', '') or ''))
    prog_names = [xl_programs.get(pc, {}).get('한글명', pc) for pc in prog_codes if pc]
    recommendations_out.append({
        'id': rid,
        'name': rec['한글명'],
        'en': rec.get('영문명', ''),
        'type': 'Recommendation',
        'recommendedProgramIds': prog_codes,
        'recommendedPrograms': prog_names,
        'targetCount': rec.get('대상인원'),
        'matchingScore': rec.get('매칭점수'),
    })

competencyGaps_out = []
for cgid, cg in xl_comp_gaps.items():
    comp = xl_comps.get(str(cg.get('갭역량', '') or ''), {})
    competencyGaps_out.append({
        'id': cgid,
        'name': cg['한글명'],
        'en': cg.get('영문명', ''),
        'type': 'CompetencyGap',
        'gapCompetencyId': str(cg.get('갭역량', '') or ''),
        'gapCompetency': comp.get('한글명', str(cg.get('갭역량', '') or '')),
        'gapSize': cg.get('갭크기'),
    })

programEnrollments_out = []
for peid, pe in xl_enrollments.items():
    prog_codes = parse_codes(str(pe.get('등록프로그램', '') or ''))
    prog_names = [xl_programs.get(pc, {}).get('한글명', pc) for pc in prog_codes if pc]
    programEnrollments_out.append({
        'id': peid,
        'name': pe['한글명'],
        'en': pe.get('영문명', ''),
        'type': 'ProgramEnrollment',
        'enrolledProgramIds': prog_codes,
        'enrolledPrograms': prog_names,
        'enrolledCount': pe.get('등록인원'),
        'completed': pe.get('이수여부'),
    })

# Summary stats
result = {
    'meta': {
        'generatedAt': '2026-04-19',
        'totalStrategies': len(strategies_out),
        'totalPolicies': len(policies_out),
        'totalBudgets': len(budgets_out),
        'totalBudgetAmount': total_strat_budget,
        'totalBudgetStr': format_won(total_strat_budget),
        'totalPrograms': len(programs_out),
        'totalCompetencies': len(comps_out),
        'totalOrganizations': len(orgs_out),
        'totalPersons': len(persons_out),
        'totalOutcomes': len(outcomes_out),
        'totalBenefits': len(benefits_out),
        'totalCompetencyAssessments': len(competencyAssessments_out),
        'totalPolicyParticipations': len(policyParticipations_out),
        'totalRecommendations': len(recommendations_out),
        'totalCompetencyGaps': len(competencyGaps_out),
        'totalProgramEnrollments': len(programEnrollments_out),
    },
    'strategies': strategies_out,
    'organizations': orgs_out,
    'policies': policies_out,
    'budgets': budgets_out,
    'programs': programs_out,
    'competencies': comps_out,
    'targetGroups': tgroups_out,
    'persons': persons_out,
    'outcomes': outcomes_out,
    'benefits': benefits_out,
    'competencyAssessments': competencyAssessments_out,
    'policyParticipations': policyParticipations_out,
    'recommendations': recommendations_out,
    'competencyGaps': competencyGaps_out,
    'programEnrollments': programEnrollments_out,
}

Path(_ONTOLOGY_JSON).parent.mkdir(parents=True, exist_ok=True)
with open(_ONTOLOGY_JSON, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

Path(_ONTOLOGY_JS).parent.mkdir(parents=True, exist_ok=True)
with open(_ONTOLOGY_JS, 'w', encoding='utf-8') as f:
    f.write('const OntologyData = ')
    json.dump(result, f, ensure_ascii=False)
    f.write(';\n')

print("Done!")
print(f"Strategies: {len(strategies_out)}")
print(f"Policies: {len(policies_out)}")
print(f"Budgets: {len(budgets_out)} (OrgBudget + PolicyBudget)")
print(f"Programs: {len(programs_out)}")
print(f"Competencies: {len(comps_out)}")
print(f"Organizations: {len(orgs_out)}")
print(f"Persons: {len(persons_out)}")
print(f"Outcomes: {len(outcomes_out)}")
print(f"Benefits: {len(benefits_out)}")
print(f"CompetencyAssessments: {len(competencyAssessments_out)}")
print(f"PolicyParticipations: {len(policyParticipations_out)}")
print(f"Recommendations: {len(recommendations_out)}")
print(f"CompetencyGaps: {len(competencyGaps_out)}")
print(f"ProgramEnrollments: {len(programEnrollments_out)}")
print(f"\nTotal budget: {format_won(total_strat_budget)}")

print("\n[Strategy → Policy / Program counts]")
for s in strategies_out:
    print(f"  {s['name'][:20]}: {s['policyCount']} policies, {s['programCount']} programs")
